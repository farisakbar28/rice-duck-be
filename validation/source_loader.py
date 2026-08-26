"""External research workbook discovery + fingerprinting (task §9/§11).

Only the explicitly configured source directory is searched -- never arbitrary
user directories. Files are fingerprinted (size, SHA-256, sheet names) BEFORE
any semantic use. Raw / clean / legacy-simulation roles are fixed and enforced
downstream; the legacy-simulation workbook may never feed R2 predictions,
input defaults, the parameter registry, or calibration.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

EMPIRICAL_SOURCE_STATUS_OK = "OK"
EMPIRICAL_SOURCE_STATUS_BLOCKED = "BLOCKED_SOURCE_FILES_MISSING"
EMPIRICAL_SOURCE_STATUS_VERSION_MISMATCH = "SOURCE_VERSION_MISMATCH"

ROLE_RAW_RECAP = "raw_recap"
ROLE_CLEAN_COHORT = "clean_cohort"
ROLE_LEGACY_SIMULATION = "legacy_simulation"

# Fixed role assignments (docs/10 section 1; task §12).
EXPECTED_SOURCES: dict[str, str] = {
    ROLE_RAW_RECAP: "Recap Data CRS Bebek.xlsx",
    ROLE_CLEAN_COHORT: "DSS_Padi_Bebek_Rekap_Bersih_v10.xlsx",
    ROLE_LEGACY_SIMULATION: "Dataset Bersih Rekap Include Hasil Simulasi Baru.xlsx",
}

# Closed Phase-5C source package. Hashes identify the reviewed workbooks;
# matching a filename alone is not sufficient evidence.
EXPECTED_SHA256: dict[str, str] = {
    ROLE_RAW_RECAP: "6b73b34a418d36cddbdf61679944eeaaf7fda312f1ee88c64476670bc0da82d1",
    ROLE_CLEAN_COHORT: "98fff237d24b6191d0d21a9e04048d54d4efd83082a61dae0da37c584405c2bd",
    ROLE_LEGACY_SIMULATION: "f5e88945196c2300d57c36b16fc5c175c24ecfbd903468cc1c553771e1e94a46",
}

ROLE_RULES: dict[str, dict[str, str]] = {
    ROLE_RAW_RECAP: {
        "allowed_use": "provenance audit / zero-vs-missing tracing only",
        "forbidden_use": "calibration, parameter derivation, prediction source",
    },
    ROLE_CLEAN_COHORT: {
        "allowed_use": "comparator cohort (36 keep + 8 excluded) only",
        "forbidden_use": "fitting, median parameters, baseline yield, calibration",
    },
    ROLE_LEGACY_SIMULATION: {
        "allowed_use": "legacy simulation audit only",
        "forbidden_use": (
            "R2 prediction source; R2 input default; parameter registry entry; "
            "calibration coefficient"
        ),
    },
}


@dataclass(frozen=True)
class SourceFile:
    role: str
    filename: str
    status: str  # PRESENT | MISSING | PRESENT_UNREADABLE_OPENPYXL_MISSING
    path: str | None = None
    size_bytes: int | None = None
    sha256: str | None = None
    sheet_names: list[str] | None = None

    @property
    def present(self) -> bool:
        return self.status == "PRESENT"

    @property
    def fingerprint_valid(self) -> bool:
        expected = EXPECTED_SHA256.get(self.role)
        return self.present and expected is not None and self.sha256 == expected

    def to_dict(self) -> dict:
        return {
            "role": self.role,
            "role_rules": ROLE_RULES[self.role],
            "filename": self.filename,
            "status": self.status,
            "path": self.path,
            "size_bytes": self.size_bytes,
            "sha256": self.sha256,
            "expected_sha256": EXPECTED_SHA256.get(self.role),
            "fingerprint_valid": self.fingerprint_valid,
            "sheet_names": self.sheet_names,
        }


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sheet_names(path: Path) -> list[str] | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def discover_sources(source_dir: Path | None) -> dict[str, SourceFile]:
    """Fingerprint expected workbooks found ONLY inside ``source_dir``."""
    results: dict[str, SourceFile] = {}
    for role, filename in EXPECTED_SOURCES.items():
        if source_dir is None or not source_dir.is_dir():
            results[role] = SourceFile(
                role=role, filename=filename, status="MISSING"
            )
            continue
        candidate = source_dir / filename
        if not candidate.is_file():
            results[role] = SourceFile(
                role=role, filename=filename, status="MISSING"
            )
            continue
        sheets = _sheet_names(candidate)
        status = "PRESENT" if sheets is not None else (
            "PRESENT_UNREADABLE_OPENPYXL_MISSING"
        )
        digest = sha256_file(candidate)
        if sheets is not None and digest != EXPECTED_SHA256[role]:
            status = "SOURCE_VERSION_MISMATCH"
        results[role] = SourceFile(
            role=role,
            filename=filename,
            status=status,
            path=str(candidate),
            size_bytes=candidate.stat().st_size,
            sha256=digest,
            sheet_names=sheets,
        )
    return results


def empirical_source_status(sources: dict[str, SourceFile]) -> str:
    """Official empirical evidence requires reviewed raw + clean fingerprints."""
    required = (sources[ROLE_RAW_RECAP], sources[ROLE_CLEAN_COHORT])
    if any(src.status == "SOURCE_VERSION_MISMATCH" for src in required):
        return EMPIRICAL_SOURCE_STATUS_VERSION_MISMATCH
    if all(src.fingerprint_valid for src in required):
        return EMPIRICAL_SOURCE_STATUS_OK
    return EMPIRICAL_SOURCE_STATUS_BLOCKED


def fingerprint_summary(sources: dict[str, SourceFile]) -> list[dict]:
    return [sources[role].to_dict() for role in EXPECTED_SOURCES]
