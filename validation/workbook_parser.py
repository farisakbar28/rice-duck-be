"""Validation-only parsers and deterministic R2 cohort reconstruction.

The raw recap is used to verify source rows and distinguish recorded values
from blanks/formulas. The clean-v10 workbook defines clean and excluded roles.
The legacy simulation workbook is parsed only into an audit inventory; none of
its values are exposed as R2 inputs, predictions, parameters, or comparators.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from app.engines.r2.normalization import normalize_cultivar_group_label
from validation.fixture_builder import (
    PRIOR_AUDIT_COUNTS,
    SUPPORTED_AGE_ASSUMPTIONS_DAYS,
    VALIDATION_ANCHOR_PLANTING_DATE,
)
from validation.source_loader import (
    ROLE_CLEAN_COHORT,
    ROLE_LEGACY_SIMULATION,
    ROLE_RAW_RECAP,
    SourceFile,
)

SOURCE_VERSION_MISMATCH = "SOURCE_VERSION_MISMATCH"
RECONSTRUCTION_OK = "RECONSTRUCTION_OK"


@dataclass(frozen=True)
class ParsedWorkbook:
    role: str
    records: tuple[dict[str, Any], ...]
    audit: dict[str, Any]


@dataclass(frozen=True)
class Reconstruction:
    status: str
    raw_records: tuple[dict[str, Any], ...]
    clean_records: tuple[dict[str, Any], ...]
    stress_records: tuple[dict[str, Any], ...]
    strict_records: tuple[dict[str, Any], ...]
    calendar_records: tuple[dict[str, Any], ...]
    counts: dict[str, int]
    mismatches: tuple[str, ...]

    def manifest(self) -> dict[str, Any]:
        return {
            "status": self.status,
            "counts": self.counts,
            "mismatches": list(self.mismatches),
            "source_rows": {
                "raw": [row["source_row"] for row in self.raw_records],
                "clean": [row["source_row"] for row in self.clean_records],
                "excluded_stress": [
                    row["source_row"] for row in self.stress_records
                ],
                "strict_supported_domain": [
                    row["source_row"] for row in self.strict_records
                ],
                "calendar_eligible": [
                    row["source_row"] for row in self.calendar_records
                ],
            },
        }


def _load_workbook(path: Path, *, data_only: bool):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency gate
        raise RuntimeError("openpyxl is required for validation workbooks") from exc
    return load_workbook(path, read_only=True, data_only=data_only)


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return None
    return None


def _positive(value: Any) -> float | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    number = float(value)
    return number if number > 0 else None


def _system(value: Any) -> tuple[str | None, str]:
    text = value.strip() if isinstance(value, str) else ""
    if text == "Jarwo 2:1":
        return "jajar_legowo", "OBSERVED"
    if text == "Tegel":
        return "tegel", "OBSERVED"
    if text == "Null(default Jarwo 2:1)":
        return "jajar_legowo", "LOCAL_DEFAULT"
    return None, "UNAVAILABLE"


def _public_variety(group_code: str | None) -> str | None:
    if group_code == "SERTANI_GROUP":
        return "sertani"
    if group_code == "INPARI_GROUP":
        return "inpari"
    return None


def _actual_provenance(
    *, clean_value: Any, raw_value: Any, raw_formula: Any
) -> str:
    if isinstance(raw_formula, str) and raw_formula.startswith("="):
        return "DERIVED_ACTUAL"
    if raw_value is None or raw_value == "":
        return "MISSING_UNKNOWN"
    if isinstance(raw_value, (int, float)) and float(raw_value) == 0:
        return "EXPLICIT_ZERO"
    if clean_value is None:
        return "MISSING_UNKNOWN"
    return "OBSERVED_VALUE"


def parse_raw_recap(path: Path) -> ParsedWorkbook:
    values_wb = _load_workbook(path, data_only=True)
    formulas_wb = _load_workbook(path, data_only=False)
    try:
        values = values_wb[values_wb.sheetnames[0]]
        formulas = formulas_wb[formulas_wb.sheetnames[0]]
        records: list[dict[str, Any]] = []
        for row_number in range(3, (values.max_row or 0) + 1):
            area = values.cell(row_number, 8).value
            # Section-heading rows do not carry numeric program area.
            if _positive(area) is None:
                continue
            records.append(
                {
                    "source_row": row_number,
                    "farmer_name_private": values.cell(row_number, 1).value,
                    "area_are": area,
                    "duck_count": values.cell(row_number, 39).value,
                    "planting_system": values.cell(row_number, 98).value,
                    "variety": values.cell(row_number, 99).value,
                    "planting_date": values.cell(row_number, 101).value,
                    "harvest_date": (
                        values.cell(row_number, 105).value
                        or values.cell(row_number, 34).value
                    ),
                    "actual_values": {
                        "yield_kg_per_are": values.cell(row_number, 31).value,
                        "paddy_price": values.cell(row_number, 32).value,
                        "duck_purchase_price": values.cell(row_number, 41).value,
                        "feed_cost": values.cell(row_number, 44).value,
                        "weeding_cost": values.cell(row_number, 87).value,
                        "fertilizer_cost": values.cell(row_number, 148).value,
                        "pesticide_cost": values.cell(row_number, 181).value,
                    },
                    "actual_formulas": {
                        "yield_kg_per_are": formulas.cell(row_number, 31).value,
                        "paddy_price": formulas.cell(row_number, 32).value,
                        "duck_purchase_price": formulas.cell(row_number, 41).value,
                        "feed_cost": formulas.cell(row_number, 44).value,
                        "weeding_cost": formulas.cell(row_number, 87).value,
                        "fertilizer_cost": formulas.cell(row_number, 148).value,
                        "pesticide_cost": formulas.cell(row_number, 181).value,
                    },
                }
            )
        return ParsedWorkbook(
            role=ROLE_RAW_RECAP,
            records=tuple(records),
            audit={"sheet": values.title, "candidate_cycle_rows": len(records)},
        )
    finally:
        values_wb.close()
        formulas_wb.close()


def _clean_row(row: tuple[Any, ...], *, excluded: bool) -> dict[str, Any]:
    source_row = int(row[0])
    area = _positive(row[7])
    ducks = int(row[8]) if _positive(row[8]) is not None else None
    system_code, system_provenance = _system(row[14])
    group = normalize_cultivar_group_label(row[13])
    group_code = group.value if group else None
    planting = _as_date(row[15]) if not excluded else None
    harvest = _as_date(row[16]) if not excluded else None
    density = (ducks / area) if area is not None and ducks is not None else None
    return {
        "source_row": source_row,
        "farmer_name_private": row[2],
        "area_are": area,
        "duck_count": ducks,
        "density_are": density,
        "cultivar_source_label": row[13],
        "cultivar_group_code": group_code,
        "rice_variety": _public_variety(group_code),
        "planting_system_source": row[14],
        "planting_system": system_code,
        "planting_system_provenance": system_provenance,
        "planting_date_observed": planting,
        "harvest_date_observed": harvest,
        "actual_yield_kg_per_are": row[15 if excluded else 17],
        "paddy_price": row[16 if excluded else 18],
        "duck_purchase_price": row[18 if excluded else 20],
        "feed_cost": row[19 if excluded else 21],
        "duck_sale_revenue": row[20 if excluded else 22],
        "weeding_cost": row[24 if excluded else 26],
        "fertilizer_cost": row[26 if excluded else 28],
        "pesticide_cost": row[28 if excluded else 30],
        "net_cost": row[30 if excluded else 32],
        "cage_cost": row[31 if excluded else 33],
        "field_notes": row[36 if excluded else 38],
        "exclusion_reason": row[37] if excluded else None,
    }


def parse_clean_v10(path: Path) -> tuple[ParsedWorkbook, ParsedWorkbook]:
    workbook = _load_workbook(path, data_only=True)
    try:
        clean_sheet = workbook["Dataset Actual Bersih"]
        excluded_sheet = workbook["Excluded Log Anomali"]
        clean = tuple(
            _clean_row(row, excluded=False)
            for row in clean_sheet.iter_rows(min_row=4, max_col=39, values_only=True)
            if row[0] is not None
        )
        excluded = tuple(
            _clean_row(row, excluded=True)
            for row in excluded_sheet.iter_rows(min_row=4, max_col=38, values_only=True)
            if row[0] is not None
        )
        return (
            ParsedWorkbook(ROLE_CLEAN_COHORT, clean, {"sheet": clean_sheet.title}),
            ParsedWorkbook("excluded_stress", excluded, {"sheet": excluded_sheet.title}),
        )
    finally:
        workbook.close()


def parse_legacy_simulation(path: Path) -> ParsedWorkbook:
    """Return workbook structure only; legacy values are intentionally hidden."""
    workbook = _load_workbook(path, data_only=False)
    try:
        sheets = [
            {
                "name": sheet.title,
                "max_row": sheet.max_row or 0,
                "max_column": sheet.max_column or 0,
            }
            for sheet in workbook.worksheets
        ]
        return ParsedWorkbook(
            ROLE_LEGACY_SIMULATION,
            (),
            {
                "role": "AUDIT_ONLY",
                "sheets": sheets,
                "values_exposed_to_r2": False,
            },
        )
    finally:
        workbook.close()


def _pseudonymize(
    rows: list[dict[str, Any]], private_map_path: Path | None
) -> None:
    names: list[str] = []
    for row in sorted(rows, key=lambda item: item["source_row"]):
        name = str(row.pop("farmer_name_private") or "").strip()
        if name not in names:
            names.append(name)
        row["farmer_cluster_id"] = f"F{names.index(name) + 1:03d}"
    if private_map_path is not None:
        private_map_path.parent.mkdir(parents=True, exist_ok=True)
        private_map_path.write_text(
            json.dumps(
                {f"F{index + 1:03d}": name for index, name in enumerate(names)},
                indent=2,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )


def reconstruct_cohorts(
    sources: dict[str, SourceFile], *, private_map_path: Path | None = None
) -> Reconstruction:
    raw_source = sources[ROLE_RAW_RECAP]
    clean_source = sources[ROLE_CLEAN_COHORT]
    if "SOURCE_VERSION_MISMATCH" in {raw_source.status, clean_source.status}:
        return Reconstruction(
            status=SOURCE_VERSION_MISMATCH,
            raw_records=(), clean_records=(), stress_records=(),
            strict_records=(), calendar_records=(), counts={},
            mismatches=("reviewed raw/clean fingerprint mismatch",),
        )
    if not raw_source.present or not clean_source.present:
        return Reconstruction(
            status="BLOCKED_SOURCE_FILES_MISSING",
            raw_records=(), clean_records=(), stress_records=(),
            strict_records=(), calendar_records=(), counts={},
            mismatches=("raw and clean-v10 are both required",),
        )

    raw = parse_raw_recap(Path(raw_source.path or ""))
    clean, excluded = parse_clean_v10(Path(clean_source.path or ""))
    raw_by_id = {row["source_row"]: row for row in raw.records}
    clean_rows = [dict(row) for row in clean.records]
    stress_rows = [dict(row) for row in excluded.records]
    all_rows = clean_rows + stress_rows
    _pseudonymize(all_rows, private_map_path)

    clean_ids = {row["source_row"] for row in clean_rows}
    stress_ids = {row["source_row"] for row in stress_rows}
    expected_raw_ids = clean_ids | stress_ids
    raw_rows = [dict(raw_by_id[row_id]) for row_id in sorted(expected_raw_ids)
                if row_id in raw_by_id]
    for row in raw_rows:
        row.pop("farmer_name_private", None)

    strict_rows = [
        row for row in clean_rows
        if row["planting_system_provenance"] == "OBSERVED"
        and row["density_are"] is not None
        and (
            row["planting_system"] == "jajar_legowo"
            and 2 <= row["density_are"] <= 4
            or row["planting_system"] == "tegel"
            and 2 <= row["density_are"] <= 3
        )
    ]
    calendar_rows = [
        row for row in clean_rows
        if row["planting_date_observed"] is not None
        and row["harvest_date_observed"] is not None
    ]

    # Attach seven-field input provenance and actual-field provenance without
    # ever exposing private farmer names.
    clean_by_source = {row["source_row"]: row for row in clean_rows}
    for source_row, row in clean_by_source.items():
        raw_row = raw_by_id[source_row]
        planting_value = (
            row["planting_date_observed"]
            or date.fromisoformat(VALIDATION_ANCHOR_PLANTING_DATE)
        )
        row["input_fields"] = {
            "land_area_are": {"value": row["area_are"], "provenance": "OBSERVED"},
            "duck_count": {"value": row["duck_count"], "provenance": "OBSERVED"},
            "planting_date": {
                "value": planting_value.isoformat(),
                "provenance": (
                    "OBSERVED" if row["planting_date_observed"] else
                    "VALIDATION_ASSUMPTION"
                ),
            },
            "planting_system": {
                "value": row["planting_system"],
                "provenance": row["planting_system_provenance"],
            },
            "rice_variety": {
                "value": row["rice_variety"],
                "provenance": "OBSERVED" if row["rice_variety"] else "UNAVAILABLE",
            },
            "duck_age_days": {
                "value": SUPPORTED_AGE_ASSUMPTIONS_DAYS[0],
                "provenance": "VALIDATION_ASSUMPTION",
            },
            "p_duck_buy": {
                "value": _positive(row["duck_purchase_price"]),
                "provenance": (
                    "OBSERVED" if _positive(row["duck_purchase_price"]) is not None
                    else "LOCAL_DEFAULT"
                ),
            },
        }
        row["actual_provenance"] = {}
        actual_key_map = {
            "yield_kg_per_are": "actual_yield_kg_per_are",
            "paddy_price": "paddy_price",
            "duck_purchase_price": "duck_purchase_price",
            "feed_cost": "feed_cost",
            "weeding_cost": "weeding_cost",
            "fertilizer_cost": "fertilizer_cost",
            "pesticide_cost": "pesticide_cost",
        }
        for raw_key, clean_key in actual_key_map.items():
            row["actual_provenance"][clean_key] = _actual_provenance(
                clean_value=row[clean_key],
                raw_value=raw_row["actual_values"][raw_key],
                raw_formula=raw_row["actual_formulas"][raw_key],
            )
        row["actual_provenance"]["duck_age_days"] = "LEGACY_IMPUTATION"
        row["actual_provenance"]["active_duration_days"] = "LEGACY_IMPUTATION"
        for key in ("net_cost", "cage_cost"):
            row["actual_provenance"][key] = (
                "DERIVED_ACTUAL" if _positive(row[key]) is not None
                else "MISSING_UNKNOWN"
            )

    counts = {
        "raw_total": len(raw_rows),
        "clean_keep": len(clean_rows),
        "excluded_stress": len(stress_rows),
        "strict_supported_domain": len(strict_rows),
        "calendar_eligible_both_dates": len(calendar_rows),
    }
    mismatches: list[str] = []
    if clean_ids & stress_ids:
        mismatches.append("clean/excluded source-row overlap")
    missing_raw = sorted(expected_raw_ids - set(raw_by_id))
    if missing_raw:
        mismatches.append(f"source rows absent from raw recap: {missing_raw}")
    for key, expected in PRIOR_AUDIT_COUNTS.items():
        if counts[key] != expected:
            mismatches.append(f"{key}: expected {expected}, reconstructed {counts[key]}")

    return Reconstruction(
        status=SOURCE_VERSION_MISMATCH if mismatches else RECONSTRUCTION_OK,
        raw_records=tuple(raw_rows),
        clean_records=tuple(clean_rows),
        stress_records=tuple(stress_rows),
        strict_records=tuple(strict_rows),
        calendar_records=tuple(calendar_rows),
        counts=counts,
        mismatches=tuple(mismatches),
    )
