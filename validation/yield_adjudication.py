"""Semantic adjudication of historical yield formulas.

This module is a validation-only correction layer.  It inspects the raw
workbook formula and its precedents before any prediction residual or metric
is calculated.  Provenance (``DERIVED_ACTUAL``) is deliberately kept separate
from the admissibility decision.
"""

from __future__ import annotations

import math
import re
from collections import OrderedDict
from pathlib import Path
from typing import Any

from validation.source_loader import (
    ROLE_CLEAN_COHORT,
    ROLE_RAW_RECAP,
    SourceFile,
)
from validation.workbook_parser import Reconstruction


YIELD_RAW_COLUMN = 31
YIELD_RAW_COLUMN_LETTER = "AE"
YIELD_CLEAN_COLUMN = 18
YIELD_CLEAN_COLUMN_LETTER = "R"
AREA_RAW_COLUMN_LETTER = "H"
HARVEST_QUANTITY_RAW_COLUMN_LETTER = "AA"

DERIVED_ACTUAL_ADMISSIBLE = "DERIVED_ACTUAL_ADMISSIBLE"
DERIVED_ACTUAL_NOT_ADMISSIBLE = "DERIVED_ACTUAL_NOT_ADMISSIBLE"
YIELD_PROVENANCE_VOCAB = (
    "OBSERVED_VALUE",
    "EXPLICIT_ZERO",
    "MISSING_UNKNOWN",
    "DERIVED_ACTUAL",
    "LEGACY_IMPUTATION",
)

YIELD_FORMULA_CLASS = (
    "ACTUAL_HARVESTED_GABAH_KG_DIVIDED_BY_PROGRAM_RICE_AREA_ARE"
)
YIELD_EXPLICIT_ZERO_POLICY = (
    "EXPLICIT_ZERO_EXCLUDED_UNLESS_ZERO_HARVEST_SEMANTICS_IS_INDEPENDENTLY_VERIFIED"
)

_CELL_REFERENCE = re.compile(
    r"(?<![A-Z0-9_])\$?([A-Z]{1,3})\$?(\d+)(?![A-Z0-9_])"
)


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _finite_number(value: Any) -> bool:
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(float(value))
    )


def _formula_references(formula: Any) -> list[str]:
    if not _is_formula(formula):
        return []
    return [f"{column}{row}" for column, row in _CELL_REFERENCE.findall(formula)]


def normalized_formula_pattern(formula: Any, source_row: int) -> str | None:
    """Replace row numbers while retaining absolute-reference markers."""
    if not _is_formula(formula):
        return None
    pattern = re.compile(rf"(\$?[A-Z]{{1,3}}\$?){source_row}\b")
    return pattern.sub(r"\1{row}", formula)


def _precedent_origin(cell_content: Any, cached_value: Any) -> str:
    if _is_formula(cell_content):
        return "FORMULA_DERIVED_ACTUAL"
    if _finite_number(cached_value):
        return "DIRECT_RECORDED_ACTUAL_VALUE"
    if cached_value in (None, ""):
        return "UNKNOWN_OR_MISSING"
    return "UNKNOWN"


def _precedent_semantics(column: str) -> dict[str, Any]:
    if column == HARVEST_QUANTITY_RAW_COLUMN_LETTER:
        return {
            "semantic_role": "NUMERATOR",
            "semantic_meaning": (
                "actual recorded gabah yield / harvested gabah quantity"
            ),
            "semantic_category": "DIRECT_FARM_MEASUREMENT_RECORDED_ACTUAL",
            "unit": "kg",
            "r2_compatibility": "compatible physical yield quantity",
        }
    if column == AREA_RAW_COLUMN_LETTER:
        return {
            "semantic_role": "DENOMINATOR",
            "semantic_meaning": "rice field area enrolled in the program",
            "semantic_category": "DIRECT_RECORDED_ACTIVE_CULTIVATED_AREA",
            "unit": "are",
            "r2_compatibility": (
                "compatible with R2 A_are / active rice-duck interaction area"
            ),
        }
    return {
        "semantic_role": "UNKNOWN",
        "semantic_meaning": "precedent column is not in the approved yield class",
        "semantic_category": "UNKNOWN",
        "unit": None,
        "r2_compatibility": "not established",
    }


def _source_summary(source: SourceFile) -> dict[str, Any]:
    """Return source identity without copying an absolute local path."""
    return {
        "role": source.role,
        "filename": source.filename,
        "status": source.status,
        "sha256": source.sha256,
        "expected_sha256": source.to_dict().get("expected_sha256"),
        "sheet_names": source.sheet_names,
    }


def _clean_sheet_rows(sheet: Any) -> dict[int, int]:
    source_to_sheet_row: dict[int, int] = {}
    for sheet_row in range(4, (sheet.max_row or 0) + 1):
        value = sheet.cell(sheet_row, 1).value
        if value is not None:
            source_to_sheet_row[int(value)] = sheet_row
    return source_to_sheet_row


def _adjudicate_record(
    *,
    source_row: int,
    clean_record: dict[str, Any],
    raw_formula_sheet: Any,
    raw_value_sheet: Any,
    clean_sheet: Any,
    clean_sheet_row: int | None,
    raw_headers: dict[str, Any],
    clean_headers: dict[str, Any],
) -> dict[str, Any]:
    source_cell = f"{YIELD_RAW_COLUMN_LETTER}{source_row}"
    formula = raw_formula_sheet.cell(source_row, YIELD_RAW_COLUMN).value
    cached_value = raw_value_sheet.cell(source_row, YIELD_RAW_COLUMN).value
    references = _formula_references(formula)
    precedents: list[dict[str, Any]] = []

    for reference in references:
        match = re.fullmatch(r"([A-Z]{1,3})(\d+)", reference)
        if match is None:  # pragma: no cover - guarded by _formula_references
            continue
        column, row_text = match.groups()
        ref_row = int(row_text)
        formula_content = raw_formula_sheet[reference].value
        raw_value = raw_value_sheet[reference].value
        semantic = _precedent_semantics(column)
        precedents.append(
            {
                "cell": reference,
                "column": column,
                "row": ref_row,
                "header": raw_headers.get(column),
                "formula": formula_content if _is_formula(formula_content) else None,
                "raw_value": raw_value,
                "origin": _precedent_origin(formula_content, raw_value),
                **semantic,
            }
        )

    normalized = normalized_formula_pattern(formula, source_row)
    canonical = normalized.replace("$", "") if normalized else None
    precedent_columns = [item["column"] for item in precedents]
    known_formula = (
        canonical == "=AA{row}/H{row}"
        and precedent_columns == [HARVEST_QUANTITY_RAW_COLUMN_LETTER, AREA_RAW_COLUMN_LETTER]
    )
    formula_class = YIELD_FORMULA_CLASS if known_formula else "UNKNOWN_YIELD_FORMULA_CLASS"

    numerator = next(
        (item for item in precedents if item["column"] == HARVEST_QUANTITY_RAW_COLUMN_LETTER),
        None,
    )
    denominator = next(
        (item for item in precedents if item["column"] == AREA_RAW_COLUMN_LETTER),
        None,
    )
    reasons: list[str] = []
    if not _is_formula(formula):
        reasons.append("RAW_YIELD_SOURCE_IS_NOT_A_FORMULA")
    if not known_formula:
        reasons.append("FORMULA_IS_NOT_APPROVED_ACTUAL_QUANTITY_DIVIDED_BY_AREA")
    if not _finite_number(cached_value):
        reasons.append("RAW_FORMULA_CACHED_VALUE_IS_NOT_FINITE_NUMERIC")
    if numerator is None or numerator["origin"] != "DIRECT_RECORDED_ACTUAL_VALUE":
        reasons.append("NUMERATOR_IS_NOT_DIRECT_RECORDED_ACTUAL_VALUE")
    if denominator is None or denominator["origin"] != "DIRECT_RECORDED_ACTUAL_VALUE":
        reasons.append("DENOMINATOR_IS_NOT_DIRECT_RECORDED_AREA_VALUE")
    if numerator is not None and not _finite_number(numerator["raw_value"]):
        reasons.append("NUMERATOR_IS_NOT_FINITE_NUMERIC")
    if denominator is not None and (
        not _finite_number(denominator["raw_value"])
        or float(denominator["raw_value"]) <= 0
    ):
        reasons.append("DENOMINATOR_IS_NOT_POSITIVE_AREA")
    if clean_record.get("actual_provenance", {}).get("actual_yield_kg_per_are") != "DERIVED_ACTUAL":
        reasons.append("PROVENANCE_IS_NOT_DERIVED_ACTUAL")

    admissible = not reasons
    clean_actual = clean_record.get("actual_yield_kg_per_are")
    clean_actual_cell = (
        f"{YIELD_CLEAN_COLUMN_LETTER}{clean_sheet_row}"
        if clean_sheet_row is not None
        else None
    )
    reason = (
        "Deterministic transformation of direct recorded actual gabah quantity "
        "(kg) divided by direct recorded program rice area (are); compatible "
        "with the R2 kg/are target."
        if admissible
        else "; ".join(reasons)
    )
    return {
        "source_row": source_row,
        "raw_yield_source_cell": source_cell,
        "raw_yield_source_header": raw_headers.get(YIELD_RAW_COLUMN_LETTER),
        "raw_formula": formula,
        "raw_formula_cached_value": cached_value,
        "formula_pattern": normalized,
        "formula_pattern_canonical": canonical,
        "formula_class": formula_class,
        "clean_yield_source_cell": clean_actual_cell,
        "clean_yield_source_header": clean_headers.get(YIELD_CLEAN_COLUMN_LETTER),
        "actual_numeric_value": clean_actual,
        "actual_provenance": clean_record.get("actual_provenance", {}).get(
            "actual_yield_kg_per_are"
        ),
        "precedents": precedents,
        "semantic_assessment": {
            "numerator": (
                "actual harvested gabah quantity, directly recorded in the raw "
                "workbook as 'Actual gabah yield (kg)'"
            ),
            "denominator": (
                "active cultivated/program rice area, directly recorded as "
                "'Rice field in program (Are)'"
            ),
            "area_type": "ACTIVE_CULTIVATED_PROGRAM_RICE_AREA",
            "area_is_total_land": False,
            "area_is_distinct_harvested_area": False,
            "area_compatible_with_r2_kg_per_are": denominator is not None
            and denominator["semantic_category"] == "DIRECT_RECORDED_ACTIVE_CULTIVATED_AREA",
            "quantity_type": "ACTUAL_HARVESTED_GABAH_QUANTITY_NOT_SALES_QUANTITY",
            "quantity_basis": (
                "Gabah basis is named, but the workbook does not identify GKP, "
                "GKG, or a moisture percentage."
            ),
            "separate_sales_fields_present": [
                "Price per kg of gabah sold",
                "Total gabah revenue",
            ],
            "moisture_basis_limitation": True,
            "forbidden_semantics_detected": [],
            "unit_transformation": "kg / are -> kg/are",
            "constant_used": None,
            "pure_deterministic_measurement_derivation": known_formula
            and all(
                item["origin"] == "DIRECT_RECORDED_ACTUAL_VALUE"
                for item in precedents
            ),
        },
        "admissibility": DERIVED_ACTUAL_ADMISSIBLE if admissible else DERIVED_ACTUAL_NOT_ADMISSIBLE,
        "derived_actual_admissibility": admissible,
        "admissibility_reason": reason,
    }


def build_yield_actual_adjudication(
    sources: dict[str, SourceFile],
    reconstruction: Reconstruction,
    *,
    scientific_target_sha: str,
    original_official_run_id: str,
    validation_harness_sha: str | None,
    protocol_source_sha: str,
) -> dict[str, Any]:
    """Audit every clean row's raw yield formula without calculating metrics."""
    raw_source = sources[ROLE_RAW_RECAP]
    clean_source = sources[ROLE_CLEAN_COHORT]
    if not raw_source.present or not raw_source.path:
        raise ValueError("raw recap source is required for yield adjudication")
    if not clean_source.present or not clean_source.path:
        raise ValueError("clean cohort source is required for yield adjudication")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency gate
        raise RuntimeError("openpyxl is required for yield adjudication") from exc

    raw_values_workbook = load_workbook(raw_source.path, read_only=True, data_only=True)
    raw_formula_workbook = load_workbook(raw_source.path, read_only=True, data_only=False)
    clean_workbook = load_workbook(clean_source.path, read_only=True, data_only=True)
    try:
        raw_values = raw_values_workbook[raw_values_workbook.sheetnames[0]]
        raw_formulas = raw_formula_workbook[raw_formula_workbook.sheetnames[0]]
        clean_sheet = clean_workbook["Dataset Actual Bersih"]
        clean_rows = _clean_sheet_rows(clean_sheet)
        raw_headers = {
            raw_formulas.cell(2, column).column_letter: raw_formulas.cell(2, column).value
            for column in range(1, (raw_formulas.max_column or 0) + 1)
        }
        clean_headers = {
            clean_sheet.cell(3, column).column_letter: clean_sheet.cell(3, column).value
            for column in range(1, (clean_sheet.max_column or 0) + 1)
        }
        clean_by_source = {
            int(row["source_row"]): row
            for row in reconstruction.clean_records
        }
        records = []
        for source_row in sorted(clean_by_source):
            records.append(
                _adjudicate_record(
                    source_row=source_row,
                    clean_record=clean_by_source[source_row],
                    raw_formula_sheet=raw_formulas,
                    raw_value_sheet=raw_values,
                    clean_sheet=clean_sheet,
                    clean_sheet_row=clean_rows.get(source_row),
                    raw_headers=raw_headers,
                    clean_headers=clean_headers,
                )
            )
    finally:
        raw_values_workbook.close()
        raw_formula_workbook.close()
        clean_workbook.close()

    pattern_groups: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for record in records:
        pattern = record["formula_pattern"] or "<NON_FORMULA>"
        group = pattern_groups.setdefault(
            pattern,
            {
                "formula_pattern": record["formula_pattern"],
                "formula_pattern_canonical": record["formula_pattern_canonical"],
                "formula_class": record["formula_class"],
                "n_rows": 0,
                "source_rows": [],
                "numerator_semantic": record["semantic_assessment"]["numerator"],
                "denominator_semantic": record["semantic_assessment"]["denominator"],
                "unit_transformation": record["semantic_assessment"]["unit_transformation"],
                "constant_used": record["semantic_assessment"]["constant_used"],
                "pure_deterministic_measurement_derivation": record["semantic_assessment"][
                    "pure_deterministic_measurement_derivation"
                ],
            },
        )
        group["n_rows"] += 1
        group["source_rows"].append(record["source_row"])

    admissible_n = sum(record["derived_actual_admissibility"] for record in records)
    not_admissible_n = len(records) - admissible_n
    provenance_counts: dict[str, int] = {
        provenance: 0 for provenance in YIELD_PROVENANCE_VOCAB
    }
    for record in records:
        provenance = record["actual_provenance"]
        provenance_counts[provenance] = provenance_counts.get(provenance, 0) + 1
    non_admissible_reasons: dict[str, int] = {}
    for record in records:
        if not record["derived_actual_admissibility"]:
            reason = record["admissibility_reason"]
            non_admissible_reasons[reason] = non_admissible_reasons.get(reason, 0) + 1

    return {
        "scientific_target_sha": scientific_target_sha,
        "original_official_run_id": original_official_run_id,
        "validation_harness_sha": validation_harness_sha,
        "protocol_source_sha": protocol_source_sha,
        "adjudication_type": "POST_EXPOSURE_PROVENANCE_SEMANTIC_ADJUDICATION",
        "metric_calculation_performed": False,
        "residuals_used_for_adjudication": False,
        "model_performance_used_for_row_selection": False,
        "admissibility_rule": {
            "derived_actual_allowed_only_when": [
                "formula is a deterministic transformation",
                "precedents are semantically compatible direct actual observations",
                "numerator is actual harvested quantity",
                "denominator is active cultivated/program rice area",
                "unit transforms to kg/are",
                "no imputation, legacy prediction, forecast, fitted coefficient, or fallback",
            ],
            "derived_actual_forbidden_classes": [
                "legacy predicted yield",
                "model simulation",
                "imputed harvest",
                "assumed production",
                "fitted coefficient",
                "historical prediction field",
                "parameter fallback",
                "forecast",
                "unsupported cross-field estimate",
            ],
            "explicit_zero_policy": YIELD_EXPLICIT_ZERO_POLICY,
            "provenance_remains": "DERIVED_ACTUAL",
        },
        "source_workbooks": {
            ROLE_RAW_RECAP: _source_summary(raw_source),
            ROLE_CLEAN_COHORT: _source_summary(clean_source),
        },
        "source_semantic_evidence": {
            "raw_yield_source_header": "Actual gabah yield (kg/are)",
            "raw_numerator_header": "Actual gabah yield (kg)",
            "raw_denominator_header": "Rice field in program (Are)",
            "clean_area_header": "A_are (Luas Program)",
            "clean_yield_header": "Actual Yield (kg/are)",
            "r2_input_field": "land_area_are / A_are",
            "r2_yield_unit": "kg/are",
            "basis_limitation": (
                "The source says gabah but does not specify GKP/GKG or moisture; "
                "the limitation is retained rather than silently harmonized."
            ),
        },
        "formula_pattern_inventory": list(pattern_groups.values()),
        "row_count": len(records),
        "admissible_derived_actual_n": admissible_n,
        "not_admissible_derived_actual_n": not_admissible_n,
        "not_admissible_reason_counts": non_admissible_reasons,
        "yield_provenance_distribution": provenance_counts,
        "records": records,
    }


def adjudication_index(adjudication: dict[str, Any] | None) -> dict[int, dict[str, Any]]:
    if not adjudication:
        return {}
    return {
        int(record["source_row"]): record
        for record in adjudication.get("records", [])
    }


def evaluate_yield_actual_eligibility(
    actual_provenance: str | None,
    actual: Any,
    *,
    derived_actual_admissibility: bool | None = None,
    explicit_zero_semantics_verified: bool = False,
) -> dict[str, Any]:
    """Apply endpoint-specific yield eligibility, independent of residuals."""
    numeric = _finite_number(actual)
    if actual_provenance == "OBSERVED_VALUE":
        return {
            "eligible": numeric,
            "derived_actual_admissibility": None,
            "reason": "OBSERVED_VALUE" if numeric else "OBSERVED_VALUE_NOT_FINITE_NUMERIC",
        }
    if actual_provenance == "DERIVED_ACTUAL":
        if derived_actual_admissibility is True and numeric:
            return {
                "eligible": True,
                "derived_actual_admissibility": True,
                "reason": DERIVED_ACTUAL_ADMISSIBLE,
            }
        return {
            "eligible": False,
            "derived_actual_admissibility": bool(derived_actual_admissibility),
            "reason": (
                "DERIVED_ACTUAL_ADJUDICATION_REQUIRED"
                if derived_actual_admissibility is not True
                else "DERIVED_ACTUAL_VALUE_NOT_FINITE_NUMERIC"
            ),
        }
    if actual_provenance == "EXPLICIT_ZERO":
        eligible = numeric and float(actual) == 0 and explicit_zero_semantics_verified
        return {
            "eligible": eligible,
            "derived_actual_admissibility": None,
            "reason": (
                "EXPLICIT_ZERO_SEMANTICALLY_VERIFIED"
                if eligible
                else YIELD_EXPLICIT_ZERO_POLICY
            ),
        }
    if actual_provenance == "LEGACY_IMPUTATION":
        reason = "LEGACY_IMPUTATION_NOT_GROUND_TRUTH"
    elif actual_provenance == "MISSING_UNKNOWN":
        reason = "MISSING_UNKNOWN_NOT_GROUND_TRUTH"
    else:
        reason = "UNSUPPORTED_YIELD_PROVENANCE"
    return {
        "eligible": False,
        "derived_actual_admissibility": None,
        "reason": reason,
    }


def yield_actual_comparator_eligible(
    actual_provenance: str | None,
    actual: Any,
    *,
    derived_actual_admissibility: bool | None = None,
    explicit_zero_semantics_verified: bool = False,
) -> bool:
    return evaluate_yield_actual_eligibility(
        actual_provenance,
        actual,
        derived_actual_admissibility=derived_actual_admissibility,
        explicit_zero_semantics_verified=explicit_zero_semantics_verified,
    )["eligible"]


__all__ = [
    "DERIVED_ACTUAL_ADMISSIBLE",
    "DERIVED_ACTUAL_NOT_ADMISSIBLE",
    "YIELD_FORMULA_CLASS",
    "YIELD_EXPLICIT_ZERO_POLICY",
    "YIELD_PROVENANCE_VOCAB",
    "adjudication_index",
    "build_yield_actual_adjudication",
    "evaluate_yield_actual_eligibility",
    "normalized_formula_pattern",
    "yield_actual_comparator_eligible",
]
