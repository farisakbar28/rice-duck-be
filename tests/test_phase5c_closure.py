"""Phase 5C evidence-boundary, parser, gate, and repository hygiene tests."""

from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook

from app.data.seed import FREEZE_ID, MODEL_VERSION, PARAMETER_REGISTRY_VERSION
from app.domain.models import LocalCultivarGroup
from app.engines.r2.normalization import normalize_cultivar_group_label
from validation.provenance import FreezeIdentity, evaluate_official_gate
from validation.source_loader import (
    ROLE_CLEAN_COHORT,
    ROLE_LEGACY_SIMULATION,
    ROLE_RAW_RECAP,
    SourceFile,
)
from validation.workbook_parser import RECONSTRUCTION_OK, reconstruct_cohorts


def test_approved_cultivar_aliases_are_exact_and_bounded() -> None:
    expected = {
        "Sertani": LocalCultivarGroup.SERTANI_GROUP,
        "Sertani 13": LocalCultivarGroup.SERTANI_GROUP,
        "Sertani a 13": LocalCultivarGroup.SERTANI_GROUP,
        "Seratih": LocalCultivarGroup.SERTANI_GROUP,
        "Inpari": LocalCultivarGroup.INPARI_GROUP,
        "Inpari 32": LocalCultivarGroup.INPARI_GROUP,
    }
    for label, group in expected.items():
        assert normalize_cultivar_group_label(label) is group
        assert normalize_cultivar_group_label(f"  {label.upper()}  ") is group
    for unsupported in ("Sertani13", "Sertani 14", "Inpari32", "Inpari 33", ""):
        assert normalize_cultivar_group_label(unsupported) is None


def _write_raw(path: Path, source_rows: list[int]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(1, 1, "header")
    sheet.cell(2, 1, "Farmer name")
    for index, source_row in enumerate(source_rows, start=1):
        sheet.cell(source_row, 1, f"Private Farmer {index % 7}")
        sheet.cell(source_row, 8, 10)
        sheet.cell(source_row, 31, f"=270/{10}")
        sheet.cell(source_row, 32, 6000)
        sheet.cell(source_row, 39, 30)
        sheet.cell(source_row, 41, 25000 if index <= 29 else None)
        sheet.cell(source_row, 44, 10000 if index <= 23 else None)
        sheet.cell(source_row, 87, None)
        sheet.cell(source_row, 98, "Jarwo 2:1")
        sheet.cell(source_row, 99, "Sertani")
        sheet.cell(source_row, 148, 1000 if index == 1 else None)
        sheet.cell(source_row, 181, 1000 if index <= 4 else None)
    workbook.save(path)


def _clean_values(source_row: int, index: int, *, excluded: bool) -> list:
    row = [None] * 39
    row[0] = source_row
    row[2] = f"Private Farmer {index % 7}"
    row[7] = 10
    row[8] = 30
    row[13] = "Sertani 13" if index % 2 else "Inpari 32"
    row[14] = "Jarwo 2:1" if excluded or index <= 17 else "Null(default Jarwo 2:1)"
    if excluded:
        row[15] = 27
        row[16] = 6000
        row[18] = 25000
        row[19] = 10000
        row[37] = "synthetic excluded condition"
    else:
        if index <= 12:
            row[15] = datetime(2025, 1, 1)
            row[16] = datetime(2025, 4, 11)
        row[17] = 27
        row[18] = 6000
        row[20] = 25000 if index <= 29 else (0 if index <= 33 else None)
        row[21] = 10000 if index <= 23 else 0
        row[26] = 0
        row[28] = 1000 if index == 1 else 0
        row[30] = 1000 if index <= 4 else 0
        row[32] = 1000 if index <= 17 else 0
        row[33] = 1000 if index <= 9 else 0
    return row


def _write_clean(path: Path, clean_ids: list[int], excluded_ids: list[int]) -> None:
    workbook = Workbook()
    protocol = workbook.active
    protocol.title = "Protokol & Klasifikasi"
    clean = workbook.create_sheet("Dataset Actual Bersih")
    excluded = workbook.create_sheet("Excluded Log Anomali")
    for column in range(1, 40):
        clean.cell(3, column, f"h{column}")
        excluded.cell(3, column, f"h{column}")
    for index, source_row in enumerate(clean_ids, start=1):
        clean.append(_clean_values(source_row, index, excluded=False))
    for index, source_row in enumerate(excluded_ids, start=1):
        excluded.append(_clean_values(source_row, index, excluded=True))
    workbook.save(path)


def test_cohorts_are_reconstructed_not_hardcoded() -> None:
    local_tmp = Path(__file__).resolve().parents[1] / "validation" / "local"
    local_tmp.mkdir(parents=True, exist_ok=True)
    temp_dir = TemporaryDirectory(dir=local_tmp)
    tmp_path = Path(temp_dir.name)
    clean_ids = list(range(4, 40))
    excluded_ids = list(range(40, 48))
    raw_path = tmp_path / "raw.xlsx"
    clean_path = tmp_path / "clean.xlsx"
    _write_raw(raw_path, clean_ids + excluded_ids)
    _write_clean(clean_path, clean_ids, excluded_ids)
    sources = {
        ROLE_RAW_RECAP: SourceFile(
            ROLE_RAW_RECAP, raw_path.name, "PRESENT", path=str(raw_path)
        ),
        ROLE_CLEAN_COHORT: SourceFile(
            ROLE_CLEAN_COHORT, clean_path.name, "PRESENT", path=str(clean_path)
        ),
        ROLE_LEGACY_SIMULATION: SourceFile(
            ROLE_LEGACY_SIMULATION, "legacy.xlsx", "MISSING"
        ),
    }
    result = reconstruct_cohorts(sources)
    assert result.status == RECONSTRUCTION_OK
    assert result.counts == {
        "raw_total": 44,
        "clean_keep": 36,
        "excluded_stress": 8,
        "strict_supported_domain": 17,
        "calendar_eligible_both_dates": 12,
    }
    assert all(row["farmer_cluster_id"].startswith("F") for row in result.clean_records)
    assert all("farmer_name_private" not in row for row in result.clean_records)
    assert result.clean_records[29]["input_fields"]["p_duck_buy"]["provenance"] == "LOCAL_DEFAULT"
    temp_dir.cleanup()


def test_official_gate_requires_sources_and_reconstruction() -> None:
    identity = FreezeIdentity(
        model_version=MODEL_VERSION,
        parameter_registry_version=PARAMETER_REGISTRY_VERSION,
        freeze_id=FREEZE_ID,
        model_frozen=True,
        freeze_effective_from="2026-08-26",
        model_commit_sha_env=None,
        app_version="test",
        history_schema_version=4,
        python_version="test",
        execution_timestamp_utc="2026-08-26T00:00:00+00:00",
    )
    blocked = evaluate_official_gate(
        identity, head="abc", tree_clean=True, tests_passed=True,
        source_discovery_executed=True,
    )
    assert "SOURCE_FINGERPRINTS_INVALID_OR_MISSING" in blocked.failed_conditions
    assert "COHORT_RECONSTRUCTION_NOT_VERIFIED" in blocked.failed_conditions
    passed = evaluate_official_gate(
        identity, head="abc", tree_clean=True, tests_passed=True,
        source_discovery_executed=True, source_fingerprints_valid=True,
        cohort_reconstruction_successful=True, source_version_mismatch=False,
    )
    assert passed.official is True


def test_private_sources_are_ignored_and_ci_is_minimal() -> None:
    root = Path(__file__).resolve().parents[1]
    assert "/penelitian/" in (root / ".gitignore").read_text(encoding="utf-8")
    workflow = (root / ".github" / "workflows" / "ci.yml").read_text(encoding="utf-8")
    assert "python -m pytest -q" in workflow
    assert "python -m compileall -q app validation" in workflow
